"""Round-trip test for the MODON CRS Excel writer/reader.

Run with: python -m pytest tests/ (or just: python tests/test_excel_io.py)
"""

import sys
import os
import io

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from src.excel_io import write_excel, read_reviewer_comments, safe_excel_text
from src.review_pipeline import Finding


def make_sample_findings():
    return [
        Finding(
            finding_id="f1", source_file="report.pdf", page_or_item="Page 34",
            section_no="5.8.5", snapshot="Settlement Exceeds Allowable Limits",
            discipline="Geotechnical",
            review_comment="Settlement predicted at 4.8cm vs 50mm allowable.",
            status="D", framework_step=11,
        ),
        Finding(
            finding_id="f2", source_file="report.pdf", page_or_item="Page 22",
            section_no="5.7", snapshot="Missing slope stability analysis",
            discipline="Geotechnical",
            review_comment="Only global stability shown, not per-segment.",
            status="C", framework_step=12,
        ),
    ]


def test_round_trip():
    findings = make_sample_findings()
    excel_bytes = write_excel(findings, job_id="job-123")
    assert len(excel_bytes) > 0

    # Simulate the user adding a reviewer comment by re-writing with one set.
    findings[0].reviewer_comment = "Agreed, please provide pile option cost."
    excel_bytes_v2 = write_excel(findings, job_id="job-123")

    job_id, comments, error = read_reviewer_comments(excel_bytes_v2)
    assert error is None, f"Unexpected error: {error}"
    assert job_id == "job-123"
    assert comments.get("f1") == "Agreed, please provide pile option cost."
    assert "f2" not in comments  # no comment was set on f2
    print("test_round_trip: OK")


def test_safe_excel_text_neutralizes_formula_prefixes():
    """The core defense against Excel/CSV formula injection: a leading
    =/+/-/@ makes Excel treat a cell as a FORMULA, not text, when opened.
    Prepending an apostrophe is the standard defusing technique - Excel
    displays the text as entered (prefix included) but never evaluates it."""
    assert safe_excel_text("=SUM(A1:A10)") == "'=SUM(A1:A10)"
    assert safe_excel_text("+32.4m elevation") == "'+32.4m elevation"
    assert safe_excel_text("-2.45m foundation level") == "'-2.45m foundation level"
    assert safe_excel_text("@mention") == "'@mention"
    assert safe_excel_text("Normal text, no prefix") == "Normal text, no prefix"
    assert safe_excel_text(None) == ""
    assert safe_excel_text(42) == "42"
    print("test_safe_excel_text_neutralizes_formula_prefixes: OK")


def test_safe_excel_text_strips_illegal_characters_and_truncates():
    """Control characters (e.g. from a messy PDF extraction or adversarial
    LLM output) raise IllegalCharacterError if written to a cell unfiltered.
    Also enforces Excel's own 32767-char hard cell limit."""
    assert safe_excel_text("before\x00after") == "beforeafter"  # null byte stripped
    assert safe_excel_text("before\x0bafter") == "beforeafter"  # vertical tab stripped
    long_text = "x" * 40000
    result = safe_excel_text(long_text)
    assert len(result) == 32767
    print("test_safe_excel_text_strips_illegal_characters_and_truncates: OK")


def test_write_excel_defuses_formula_injection_end_to_end():
    """Not just the helper function in isolation - verify a Finding with
    adversarial/naturally-formula-prefixed content actually comes out of the
    real write_excel() pipeline as a STRING cell, not a FORMULA cell.
    Geotechnical elevation notation (leading +/-) makes this a routine,
    non-malicious case for this domain, not just an adversarial one."""
    findings = [
        Finding(
            finding_id="f1", source_file="report.pdf", page_or_item="Page 5",
            section_no="2", snapshot="+32.4m elevation discrepancy",
            discipline="Geotechnical",
            review_comment="=cmd|'/c calc'!A1 disguised as a finding, or just a coincidental leading equals sign",
            status="D", framework_step=1,
        ),
    ]
    excel_bytes = write_excel(findings, job_id="job-123")

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    snapshot_cell = ws.cell(2, 4)   # column D: Snapshot
    comment_cell = ws.cell(2, 7)    # column G: Review comments

    assert snapshot_cell.data_type == "s", (
        f"snapshot cell must be plain text (data_type 's'), got {snapshot_cell.data_type!r} "
        f"- a formula-typed cell here means the injection defense failed")
    assert snapshot_cell.value == "'+32.4m elevation discrepancy"

    assert comment_cell.data_type == "s", (
        f"comment cell must be plain text (data_type 's'), got {comment_cell.data_type!r} "
        f"- a formula-typed cell here means the injection defense failed")
    assert comment_cell.value.startswith("'=")
    print("test_write_excel_defuses_formula_injection_end_to_end: OK")


def test_rejects_wrong_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Not", "The", "Right", "Columns"])
    buf = io.BytesIO()
    wb.save(buf)

    job_id, comments, error = read_reviewer_comments(buf.getvalue())
    assert error is not None
    assert job_id is None
    print("test_rejects_wrong_template: OK")


if __name__ == "__main__":
    test_round_trip()
    test_safe_excel_text_neutralizes_formula_prefixes()
    test_safe_excel_text_strips_illegal_characters_and_truncates()
    test_write_excel_defuses_formula_injection_end_to_end()
    test_rejects_wrong_template()
    print("All tests passed.")
