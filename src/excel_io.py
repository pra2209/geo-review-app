"""Excel I/O matching the real MODON CRS template columns (confirmed from
`Comments_Modon - Template.xlsx`), plus app-added tracking:

  A Sl. No.  B Item/Page No  C Section No.  D Snapshot  E Department/Discipline
  F Comments made by  G Review comments  H Initial Status
  I Reviewer Comments        <- app addition, visible, user fills between iterations
  J Finding ID (hidden)      <- app addition, for round-trip tracking across iterations
  K Job ID (hidden)          <- app addition, so a re-uploaded file self-identifies its job

Row 1 = headers (matches the official template exactly in columns A-H so the
output is submission-ready). A disclaimer line is appended as a footer row
below the data, not inserted above the header, to avoid disturbing the
expected header position.

Known MVP simplification: rows the user adds by hand in Excel (no Finding ID)
are not picked up on re-upload - only reviewer_comment edits against existing
rows are read back. Documented in README as a v2 improvement.
"""

import io
from typing import Optional

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.review_pipeline import Finding

HEADERS = [
    "Sl. No.", "Item/Page No", "Section No.", "Snapshot", "Department/Discipline",
    "Comments made by", "Review comments", "Initial Status",
]
EXTRA_HEADERS = ["Reviewer Comments", "Finding ID", "Job ID"]

STATUS_COLORS = {"D": "FF0000", "C": "FFA500", "B": "FFFF00", "A": "00B050", "E": "FFFFFF"}

DISCLAIMER = ("AI-assisted draft review - requires verification by a licensed geotechnical "
              "engineer before use in formal submissions.")

# Excel/CSV formula injection: a cell whose text starts with one of these is
# interpreted as a FORMULA when the file is opened, not literal text. Every
# string written here originates from either the LLM (report content can
# itself be adversarial - prompt injection is explicitly a threat model for
# this pipeline - or the model can independently emit a leading +/-/=/@) or
# a re-uploaded reviewer comment (untrusted the moment this workbook might
# be forwarded to someone else). This is NOT a theoretical risk for this
# domain specifically: geotechnical elevation notation uses leading +/- as a
# matter of routine ("+32.4m", "-2.45m" appear constantly in real reports),
# so genuine, non-malicious content will trip this if unhandled.
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_EXCEL_CELL_CHAR_LIMIT = 32767  # Excel's own hard per-cell limit


def safe_excel_text(value: object) -> str:
    """Neutralizes formula injection (prefixes a defusing apostrophe - Excel
    then displays the text as entered, formula-looking prefix included, but
    never evaluates it) and strips characters openpyxl cannot write at all
    (control characters raise IllegalCharacterError otherwise). Apply to
    EVERY untrusted string before it reaches a cell - LLM output and
    re-uploaded reviewer comments alike."""
    text = "" if value is None else str(value)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    if text.startswith(_FORMULA_PREFIXES):
        text = "'" + text
    return text[:_EXCEL_CELL_CHAR_LIMIT]


def write_excel(findings: list[Finding], job_id: str, comments_by: str = "System Review") -> bytes:
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active
    ws.title = "Sheet1"

    all_headers = HEADERS + EXTRA_HEADERS
    for col, header in enumerate(all_headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(name="Arial", size=10, bold=True)

    for row_idx, finding in enumerate(findings, start=2):
        finding.sl_no = row_idx - 1
        values = [
            finding.sl_no,  # int, not untrusted text - left as-is so it stays a real number in the sheet
            safe_excel_text(finding.page_or_item),
            safe_excel_text(finding.section_no),
            safe_excel_text(finding.snapshot),
            safe_excel_text(finding.discipline),
            safe_excel_text(comments_by),
            safe_excel_text(finding.review_comment),
            safe_excel_text(finding.status),
            safe_excel_text(finding.reviewer_comment),
            safe_excel_text(finding.finding_id),
            safe_excel_text(job_id),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col, value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        status_cell = ws.cell(row_idx, 8)
        color = STATUS_COLORS.get(finding.status, "FFFFFF")
        status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Hidden tracking columns (Finding ID, Job ID)
    ws.column_dimensions["J"].hidden = True
    ws.column_dimensions["K"].hidden = True

    # Visible column widths
    widths = {"A": 8, "B": 14, "C": 12, "D": 32, "E": 16, "F": 16, "G": 60, "H": 12, "I": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    disclaimer_row = len(findings) + 3
    ws.cell(disclaimer_row, 1, DISCLAIMER).font = Font(name="Arial", size=9, italic=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def read_reviewer_comments(file_bytes: bytes) -> tuple[Optional[str], dict[str, str], Optional[str]]:
    """Parse a re-uploaded annotated Excel.

    Returns (job_id, {finding_id: reviewer_comment}, error_message).
    error_message is set (and the other two are None/{}) if the file doesn't
    look like one this app generated.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception as exc:  # noqa: BLE001
        return None, {}, f"Could not open the uploaded file as Excel: {exc}"

    header_row = [ws.cell(1, c).value for c in range(1, 12)]
    expected = HEADERS + EXTRA_HEADERS
    if header_row[:len(HEADERS)] != HEADERS:
        return None, {}, (
            "This file's columns don't match the expected review template. "
            "Please upload the Excel that was generated by this app (don't rename/reorder columns)."
        )

    job_id = None
    comments: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_col=11):
        finding_id = row[9].value  # column J
        row_job_id = row[10].value  # column K
        if not finding_id:
            continue
        if job_id is None:
            job_id = row_job_id
        reviewer_comment = row[8].value  # column I
        if reviewer_comment:
            comments[str(finding_id)] = str(reviewer_comment)

    return job_id, comments, None
